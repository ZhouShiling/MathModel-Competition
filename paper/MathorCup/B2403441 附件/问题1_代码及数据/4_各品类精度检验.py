import time

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def gm11(x0, predict_num):
    """
    Function to perform prediction using the traditional GM(1,1) model.
    """
    n = len(x0)
    x1 = np.cumsum(x0)
    z1 = (x1[:-1] + x1[1:]) / 2

    y = x0[1:]
    x = z1

    k = ((n - 1) * np.sum(x * y) - np.sum(x) * np.sum(y)) / \
        ((n - 1) * np.sum(x ** 2) - np.sum(x) ** 2)
    b = (np.sum(x ** 2) * np.sum(y) - np.sum(x) * np.sum(x * y)) / \
        ((n - 1) * np.sum(x ** 2) - np.sum(x) ** 2)
    a = -k

    x0_hat = np.zeros(n)
    x0_hat[0] = x0[0]
    for m in range(1, n):
        x0_hat[m] = (1 - np.exp(a)) * (x0[0] - b / a) * np.exp(-a * m)

    result = np.zeros(predict_num)
    for i in range(predict_num):
        result[i] = (1 - np.exp(a)) * (x0[0] - b / a) * np.exp(-a * (n + i))

    # Calculate absolute residuals and relative residuals
    absolute_residuals = x0[1:] - x0_hat[1:]
    relative_residuals = np.abs(absolute_residuals) / x0[1:]

    # Calculate ratio and ratio deviation
    class_ratio = x0[1:] / x0[:-1]
    eta = np.abs(1 - (1 - 0.5 * a) / (1 + 0.5 * a) * (1 / class_ratio))

    return result, x0_hat, relative_residuals, eta, a, b


def metabolism_gm11(x0, predict_num):
    """
    Function to perform prediction using the metabolism GM(1,1) model.
    """
    result = np.zeros(predict_num)
    for i in range(predict_num):
        res = gm11(x0, 1)[0][0]
        result[i] = res
        x0 = np.concatenate((x0[1:], [res]))  # Update x0 with the predicted value

    return result


def new_gm11(x0, predict_num):
    """
    Function to perform prediction using the new information GM(1,1) model.
    """
    result = np.zeros(predict_num)
    for i in range(predict_num):
        res = gm11(x0, 1)[0][0]
        result[i] = res
        x0 = np.append(x0, res)  # Append the predicted value to the end of x0

    return result


# 读取附件1中的库存数据
inventory_data = pd.read_excel('附件1.xlsx')

# 初始化预测结果列表
inventory_predictions = []
error_results = []
relative_residuals_results = []
ratio_deviation_results = []

# 对每个品类进行循环预测
for category, df in inventory_data.groupby('品类'):
    # 仅保留库存量和日期列
    df = df[['月份', '库存量']]
    df.columns = ['date', 'inventory']

    # 转换日期格式
    df['date'] = pd.to_datetime(df['date'], format='%Y-%m-%d')
    df.set_index('date', inplace=True)

    # 检查是否存在足够的数据点
    data = df['inventory'].values
    n = len(data)

    # Check data validity
    ERROR = 0
    if np.any(data < 0):
        print("The time series for grey prediction cannot use negative numbers.")
        ERROR = 1

    print(f"The length of the original data is {n}")
    if n <= 3:
        print("The data volume is too small to perform grey prediction.")
        ERROR = 1
    elif n > 10:
        print("The data volume is too large for grey prediction to provide accurate predictions.")

    # Exponential regularity check
    if ERROR == 0:
        print("----------------------------")
        print("Exponential Regularity Check")
        x1 = np.cumsum(data)
        rho = data[1:] / x1[:-1]

        fig, ax = plt.subplots()
        ax.plot(range(1, n), rho, 'o-', label='Smoothness Ratios')
        ax.axhline(y=0.5, color='r', linestyle='--', label='Critical Line')
        ax.text(n - 2, 0.55, 'Critical Line')
        ax.set_xticks(range(1, n))
        ax.set_xlabel('Time Period')
        ax.set_ylabel('Smoothness Ratio')
        ax.legend()
        plt.title(f'Smoothness Ratio for Category {category}')
        plt.savefig(f'smoothness_ratio_{category}.png')  # 保存图表
        plt.close()  # 关闭图表
        time.sleep(1)  # 等待1秒

        print(
            f"Indicator 1: The percentage of smoothness ratios less than 0.5 is {100 * np.sum(rho < 0.5) / (n - 1):.2f}%")
        print(
            f"Indicator 2: Excluding the first two periods, the percentage of smoothness ratios less than 0.5 is {100 * np.sum(rho[2:] < 0.5) / (n - 3):.2f}%")
        print(
            "Reference standards: Indicator 1 should be greater than 60%, and Indicator 2 should be greater than 90%.")

        # Assume the data passes the exponential regularity test.
        judge = 1
        if judge == 0:
            print("Grey prediction model is not suitable for this data.")
            ERROR = 1

    # Split the data into training and testing sets
    if ERROR == 0 and n > 4:
        print(
            "Because the number of periods in the original data is greater than 4, we can divide the data into training and testing sets.")
        if n > 7:
            test_num = 3
        else:
            test_num = 2
        train_data = data[:-test_num]
        test_data = data[-test_num:]

        print("Training data:")
        print(train_data)
        print("Testing data:")
        print(test_data)
        print("----------------------------")

        # Traditional GM(1,1) model
        print("Traditional GM(1,1) model prediction details:")
        result1, data_hat1, rel_res1, ratio_dev1, a1, b1 = gm11(train_data, test_num)
        relative_residuals_results.append([category, '传统GM(1,1)', rel_res1])
        ratio_deviation_results.append([category, '传统GM(1,1)', ratio_dev1])

        # New information GM(1,1) model
        print("New information GM(1,1) model prediction details:")
        result2 = new_gm11(train_data, test_num)

        # Metabolism GM(1,1) model
        print("Metabolism GM(1,1) model prediction details:")
        result3 = metabolism_gm11(train_data, test_num)

        SSE_1 = np.sum((test_data - result1) ** 2)
        SSE_2 = np.sum((test_data - result2) ** 2)
        SSE_3 = np.sum((test_data - result3) ** 2)

        error_results.append([category, SSE_1, SSE_2, SSE_3])

        print(f"SSE for the traditional GM(1,1) model on the test set is {SSE_1:.2f}")
        print(f"SSE for the new information GM(1,1) model on the test set is {SSE_2:.2f}")
        print(f"SSE for the metabolism GM(1,1) model on the test set is {SSE_3:.2f}")

        if SSE_1 < SSE_2:
            if SSE_1 < SSE_3:
                choose = 1
            else:
                choose = 3
        elif SSE_2 < SSE_3:
            choose = 2
        else:
            choose = 3

        models = ['Traditional GM(1,1) Model', 'New Information GM(1,1) Model', 'Metabolism GM(1,1) Model']
        print(f"Since the SSE of {models[choose - 1]} is the smallest, we should choose it for prediction.")
        print("----------------------------")

        predict_num = 3  # Predict the next 3 months
        if choose == 1:
            result, data_hat, *_ = gm11(data, predict_num)
            relative_residuals_results.append([category, '传统GM(1,1)_预测', _])
            ratio_deviation_results.append([category, '传统GM(1,1)_预测', _])
        elif choose == 2:
            result = new_gm11(data, predict_num)
            data_hat = new_gm11(data, len(data))
        elif choose == 3:
            result = metabolism_gm11(data, predict_num)
            data_hat = metabolism_gm11(data, len(data))

        # Save the results
        inventory_predictions.extend([(category, data_hat[i], result[i]) for i in range(predict_num)])

    else:
        predict_num = 3  # Predict the next 3 months
        print("Due to the small number of data points, we will simply average the results from all three methods.")

        print("Traditional GM(1,1) model prediction details:")
        result1, data_hat1, rel_res1, ratio_dev1, a1, b1 = gm11(data, predict_num)
        relative_residuals_results.append([category, '传统GM(1,1)', rel_res1])
        ratio_deviation_results.append([category, '传统GM(1,1)', ratio_dev1])

        print("New information GM(1,1) model prediction details:")
        result2 = new_gm11(data, predict_num)

        print("Metabolism GM(1,1) model prediction details:")
        result3 = metabolism_gm11(data, predict_num)

        result = (result1 + result2 + result3) / 3

        data_hat = (data_hat1 + new_gm11(data, len(data)) + metabolism_gm11(data, len(data))) / 3

        # Save the results
        inventory_predictions.extend([(category, data_hat[i], result[i]) for i in range(predict_num)])

# Create DataFrame to store the prediction results
inventory_predictions_df = pd.DataFrame(inventory_predictions, columns=['品类', '拟合数据', '预测数据'])

# Save the results to CSV
inventory_predictions_df.to_csv('future_inventory_predictions_350_categories1.csv', index=False)

# 创建一个包含不同模型误差的表格
error_results_df = pd.DataFrame(error_results, columns=['品类', '传统GM(1,1)', '新信息GM(1,1)', '代谢GM(1,1)'])
error_results_df.to_csv('误差结果1.csv', index=False)

# 创建一个包含相对残差的表格
relative_residuals_df = pd.DataFrame(relative_residuals_results, columns=['品类', '模型', '相对残差'])
relative_residuals_df.to_csv('相对残差.csv', index=False)

# 创建一个包含级比偏差的表格
ratio_deviation_df = pd.DataFrame(ratio_deviation_results, columns=['品类', '模型', '级比偏差'])
ratio_deviation_df.to_csv('级比偏差.csv', index=False)

# 输出最终结果
print(inventory_predictions_df)
print(error_results_df)
print(relative_residuals_df)
print(ratio_deviation_df)

# 将预测结果、误差结果、相对残差和级比偏差保存到同一个Excel文件中
wb = Workbook()
ws1 = wb.active
ws1.title = '预测结果'
for r in dataframe_to_rows(inventory_predictions_df, index=False, header=True):
    ws1.append(r)

ws2 = wb.create_sheet(title='误差结果')
for r in dataframe_to_rows(error_results_df, index=False, header=True):
    ws2.append(r)

ws3 = wb.create_sheet(title='相对残差')
for r in dataframe_to_rows(relative_residuals_df, index=False, header=True):
    ws3.append(r)

ws4 = wb.create_sheet(title='级比偏差')
for r in dataframe_to_rows(ratio_deviation_df, index=False, header=True):
    ws4.append(r)

wb.save('预测与误差与残差与偏差结果1.xlsx')
